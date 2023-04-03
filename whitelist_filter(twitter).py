import openpyxl
from selenium import webdriver
import time

file_path = 'file_path_local'
sheet_name = 'sheet_name'
wb = openpyxl.load_workbook(file_path)
ws = wb[sheet_name]

driver = webdriver.Chrome()

row_num = 2
counter = 0
while ws.cell(row=row_num, column=1).value is not None:
    handle = ws.cell(row=row_num, column=1).value
    url = f'https://twitter.com/{handle}'
    driver.get(url)
    time.sleep(5) 
    try:
        if ('此账号不存在' in driver.page_source) or ('账号已被冻结' in driver.page_source):
            print(f'{handle} is not a valid Twitter handle')
            ws.cell(row=row_num, column=2).value = "No"
        else:
            print(f'{handle} is a valid Twitter handle')
            ws.cell(row=row_num, column=2).value = "Yes"
    except:
        print(f'Unable to access {handle}\'s profile')
        ws.cell(row=row_num, column=2).value = "No"
    row_num += 1
    counter += 1
    if counter == 10:
        counter = 0
        wb.save(file_path)

wb.save(file_path)

driver.quit()
